import pdfplumber, re, os, json,sys
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from plantilla import generar_plantilla

# Busca todos los PDFs en subcarpetas de la carpeta principal
def buscar_pdfs_en_carpeta(carpeta_principal):
    pdfs_encontrados = []
    
    for subcarpeta in carpeta_principal.iterdir():
        if subcarpeta.is_dir():
            for archivo in subcarpeta.iterdir():
                if archivo.is_file() and archivo.suffix.lower() == '.pdf':
                    pdfs_encontrados.append({
                        'ruta': archivo,
                        'subcarpeta': subcarpeta,
                        'nombre_archivo': archivo.name
                    })
    
    return pdfs_encontrados

# Extrae datos de documentos usando expresiones regulares del PDF
def extraer_datos_del_pdf(archivo_pdf):
    tipos, documentos, nombres = [], [], []
    
    with pdfplumber.open(archivo_pdf) as pdf:
        for page_num in range(len(pdf.pages)):
            texto = pdf.pages[page_num].extract_text()
            if not texto:
                continue

            patron = re.compile(
                r'(\d+)\.\s+([A-Z0-9-]+)\s+([^\n]+?)\s+POR\s+CERTIFICAR\s+APROBADO'
            )

            matches = patron.findall(texto)

            if not matches:
                continue
                
            for match in matches:
                tipo_doc_completo = match[1]
                nombre = match[2].strip()
                nombre = re.sub(r'[^A-ZÁÉÍÓÚÜÑ\s]', '', nombre.upper())
                nombre = re.sub(r'\s+', ' ', nombre).strip()
                
                if '-' in tipo_doc_completo:
                    tipo, documento = tipo_doc_completo.split('-', 1)
                else:
                    tipo = ""
                    documento = tipo_doc_completo
                
                tipos.append(tipo)
                documentos.append(documento)
                nombres.append(nombre)

    if len(documentos) == 0:
        raise ValueError("No se pudo extraer el listado. El formato del PDF puede estar mal elaborado o no seguir la estructura esperada.")

    return tipos, documentos, nombres

# Llena la plantilla Excel con los datos extraídos del PDF
def llenar_plantilla_con_datos(tipos, documentos, nombres, archivo_salida):

    archivo_plantilla_temp = generar_plantilla()
    
    wb = load_workbook(archivo_plantilla_temp)
    ws = wb.active
    
    for i, (tipo, documento, nombre) in enumerate(zip(tipos, documentos, nombres), start=2):
        ws[f'A{i}'] = tipo
        ws[f'B{i}'] = documento
        ws[f'C{i}'] = nombre
    
    wb.save(archivo_salida)
    
    try:
        os.unlink(archivo_plantilla_temp)
    except:
        pass

# Función principal que procesa todos los PDFs y genera reporte de resultados
def procesar_carpeta_principal_api(ruta_carpeta):
    resultado = {
        "procesados_exitosos": 0,
        "procesados_con_error": 0,
        "total_pdfs_encontrados": 0,
        "archivos_procesados": [],
        "errores": []
    }
    
    try:
        carpeta_principal = Path(ruta_carpeta)
        
        if not carpeta_principal.exists() or not carpeta_principal.is_dir():
            raise ValueError("La carpeta no existe o la ruta es incorrecta")
        
        pdfs_encontrados = buscar_pdfs_en_carpeta(carpeta_principal)
        resultado["total_pdfs_encontrados"] = len(pdfs_encontrados)
        
        if not pdfs_encontrados:
            resultado["message"] = "No se encontraron archivos PDF en las subcarpetas."
            return resultado
        
        for pdf_info in pdfs_encontrados:
            try:
                tipos, documentos, nombres = extraer_datos_del_pdf(pdf_info['ruta'])
                
                archivo_salida = pdf_info['subcarpeta'] / "plantilla.xlsx"
                
                llenar_plantilla_con_datos(tipos, documentos, nombres, archivo_salida)
                
                resultado["procesados_exitosos"] += 1
                resultado["archivos_procesados"].append({
                    "pdf": pdf_info['nombre_archivo'],
                    "subcarpeta": str(pdf_info['subcarpeta']),
                    "excel_generado": str(archivo_salida),
                    "registros_procesados": len(documentos)
                })
                
            except Exception as e:
                resultado["procesados_con_error"] += 1
                resultado["errores"].append({
                    "pdf": pdf_info['nombre_archivo'],
                    "subcarpeta": str(pdf_info['subcarpeta']),
                    "error": str(e)
                })
                continue
        
        if resultado["procesados_exitosos"] > 0:
            resultado["message"] = f"Proceso completado. {resultado['procesados_exitosos']} PDFs procesados exitosamente."
        else:
            resultado["message"] = "No se pudo procesar ningún PDF."
        
        return resultado
        
    except Exception as e:
        resultado["error"] = str(e)
        return resultado

# Interfaz de línea de comandos para interacción con el usuario
def procesar_carpeta_principal_cli():
    print("=" * 60)
    print("      EXTRACTOR MASIVO DE DATOS PDF A PLANTILLA EXCEL")
    print("=" * 60)
    
    try:
        ruta_carpeta = input("Por favor, ingresa la ruta completa de la carpeta principal: ").strip()
        ruta_carpeta = ruta_carpeta.strip('"\'')
        
        if not os.path.exists(ruta_carpeta) or not os.path.isdir(ruta_carpeta):
            print("Error: La carpeta no existe o la ruta es incorrecta.")
            return
        
        resultado = procesar_carpeta_principal_api(ruta_carpeta)
        
        print(f"\nTotal de PDFs encontrados: {resultado['total_pdfs_encontrados']}")
        
        for archivo in resultado["archivos_procesados"]:
            print(f"{archivo['pdf']} -> {archivo['excel_generado']} ({archivo['registros_procesados']} registros)")
        
        for error in resultado["errores"]:
            print(f"❌ {error['pdf']}: {error['error']}")
        
        print(f"\nRESUMEN:")
        print(f"PDFs procesados exitosamente: {resultado['procesados_exitosos']}")
        print(f"PDFs con errores: {resultado['procesados_con_error']}")
        
        if resultado["procesados_exitosos"] > 0:
            print(f"\nLos archivos Excel se guardaron en sus respectivas subcarpetas")
            print("   con el nombre: plantilla.xlsx")
        
    except Exception as e:
        print(f"\nError durante el proceso: {str(e)}")
    
    input("\nPresiona Enter para salir...")

# Punto de entrada principal - decide entre modo CLI o API
if __name__ == "__main__":
    if len(sys.argv) == 1:
        procesar_carpeta_principal_cli()
    elif len(sys.argv) == 2:
        ruta_carpeta = sys.argv[1]
        resultado = procesar_carpeta_principal_api(ruta_carpeta)
        print(json.dumps(resultado, ensure_ascii=False, indent=2))
    else:
        print("Uso: python ExtraerInfAprendiz.py [ruta_carpeta]")
        print("Si no se proporciona ruta, se ejecuta en modo interactivo")