import pandas as pd
import tempfile
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

def generar_plantilla():
    # Crear la plantilla
    data = { 'TIPO DE DOCUMENTO': [], 'NUMERO DE DOCUMENTO': [], 'NOMBRES Y APELLIDOS': [], 'DIA': [], 'MES': [], 'AÑO': [] }
    df = pd.DataFrame(data)

    # Crear archivo temporal
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        nombre_archivo = tmp.name
    
    # Guardar la plantilla en el archivo temporal
    df.to_excel(nombre_archivo, index=False, engine='openpyxl')
    
    # Ajustar el ancho de las columnas y agregar validaciones
    ajustar_ancho(nombre_archivo)
    agregar_validacion_datos(nombre_archivo)
    
    print(f"Plantilla generada temporalmente en: {nombre_archivo}")
    return nombre_archivo

def ajustar_ancho(nombre_archivo):
    # Abrir el archivo Excel para ajustar el ancho de las columnas
    wb = load_workbook(nombre_archivo)
    ws = wb.active

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Obtener la letra de la columna
        column_name = column[0].value  # Obtener el nombre de la columna
        
        for cell in column:
            try:
                # Calcular el largo máximo del contenido de la celda
                max_length = max(max_length, len(str(cell.value) if cell.value else ""))
            except Exception as e:
                print(f"Error ajustando ancho: {e}")
        
        # Configurar un ancho específico para la columna "NOMBRE Y APELLIDO"
        if column_name == "NOMBRES Y APELLIDOS":
            adjusted_width = 40  # Ancho específico
        else:
            adjusted_width = max_length + 10  # Ancho calculado dinámicamente
        
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar los cambios en el archivo Excel
    wb.save(nombre_archivo)

def agregar_validacion_datos(nombre_archivo):
    # Abrir el archivo Excel para agregar validación de datos
    wb = load_workbook(nombre_archivo)
    ws = wb.active
    
    # Obtener el año actual
    año_actual = datetime.now().year
    
    # 1. VALIDACIÓN PARA TIPO DE DOCUMENTO (Columna A)
    dv_tipo_doc = DataValidation(
        type="list",
        formula1='"CC,TI,CE,PPT"',
        allow_blank=True
    )
    dv_tipo_doc.error = 'Debe seleccionar uno de los valores válidos: CC, TI, CE, PPT'
    dv_tipo_doc.errorTitle = 'Entrada inválida'
    dv_tipo_doc.prompt = 'Seleccione un tipo de documento válido: CC, TI, CE, PPT'
    dv_tipo_doc.promptTitle = 'Tipo de Documento'
    dv_tipo_doc.showErrorMessage = True
    dv_tipo_doc.errorStyle = 'stop'
    dv_tipo_doc.add('A2:A1000')
    ws.add_data_validation(dv_tipo_doc)
    
    # 2. VALIDACIÓN PARA DIA (Columna D) - Solo números del 1 al 31
    dv_dia = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=31,
        allow_blank=True
    )
    dv_dia.error = 'El día debe ser un número entre 1 y 31'
    dv_dia.errorTitle = 'Día inválido'
    dv_dia.prompt = 'Ingrese un día válido (1-31)'
    dv_dia.promptTitle = 'Día'
    dv_dia.showErrorMessage = True
    dv_dia.errorStyle = 'stop'
    dv_dia.add('D2:D1000')
    ws.add_data_validation(dv_dia)
    
    # 3. VALIDACIÓN PARA MES (Columna E) - Lista de meses en mayúsculas
    dv_mes = DataValidation(
        type="list",
        formula1='"ENERO,FEBRERO,MARZO,ABRIL,MAYO,JUNIO,JULIO,AGOSTO,SEPTIEMBRE,OCTUBRE,NOVIEMBRE,DICIEMBRE"',
        allow_blank=True
    )
    dv_mes.error = 'Debe seleccionar un mes válido en mayúsculas'
    dv_mes.errorTitle = 'Mes inválido'
    dv_mes.prompt = 'Seleccione un mes válido'
    dv_mes.promptTitle = 'Mes'
    dv_mes.showErrorMessage = True
    dv_mes.errorStyle = 'stop'
    dv_mes.add('E2:E1000')
    ws.add_data_validation(dv_mes)
    
    # 4. VALIDACIÓN PARA AÑO (Columna F) - Desde 1900 hasta el año actual
    dv_año = DataValidation(
        type="whole",
        operator="between",
        formula1=1900,
        formula2=año_actual,
        allow_blank=True
    )
    dv_año.error = f'El año debe estar entre 1900 y {año_actual} (no se permiten años futuros)'
    dv_año.errorTitle = 'Año inválido'
    dv_año.prompt = f'Ingrese un año válido (1900-{año_actual})'
    dv_año.promptTitle = 'Año'
    dv_año.showErrorMessage = True
    dv_año.errorStyle = 'stop'
    dv_año.add('F2:F1000')
    ws.add_data_validation(dv_año)
    
    # Guardar los cambios
    wb.save(nombre_archivo)

if __name__ == "__main__":
    generar_plantilla()