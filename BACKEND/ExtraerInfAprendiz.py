import pdfplumber
import re
import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook
import json
import sys

# Importar la función del otro componente
from plantilla import generar_plantilla

def buscar_pdfs_en_carpeta(carpeta_principal):
    """
    Busca todos los PDFs en las subcarpetas de la carpeta principal
    """
    pdfs_encontrados = []
    
    # Recorrer todas las subcarpetas
    for subcarpeta in carpeta_principal.iterdir():
        if subcarpeta.is_dir():
            # Buscar PDFs en esta subcarpeta
            for archivo in subcarpeta.iterdir():
                if archivo.is_file() and archivo.suffix.lower() == '.pdf':
                    pdfs_encontrados.append({
                        'ruta': archivo,
                        'subcarpeta': subcarpeta,
                        'nombre_archivo': archivo.name
                    })
    
    return pdfs_encontrados

def extraer_datos_del_pdf(archivo_pdf):
    """
    Extrae todos los datos del PDF sin preguntar páginas
    """
    # Listas para almacenar todos los datos
    tipos, documentos, nombres = [], [], []
    
    with pdfplumber.open(archivo_pdf) as pdf:
        for page_num in range(len(pdf.pages)):
            # Extraer texto con configuración mejorada
            texto = pdf.pages[page_num].extract_text()
            if not texto:
                continue

            # Patrón más flexible que no depende de caracteres especiales
            # Buscar el patrón numérico + tipo-doc + nombre hasta "POR CERTIFICAR"
            patron = re.compile(
                r'(\d+)\.\s+([A-Z0-9-]+)\s+([^\n]+?)\s+POR\s+CERTIFICAR\s+APROBADO'
            )

            matches = patron.findall(texto)
            
            # Si no encontramos ningún registro en esta página, continuar
            if not matches:
                continue
                
            for match in matches:
                tipo_doc_completo = match[1]
                nombre = match[2].strip()
                
                # Limpiar el nombre - puede contener caracteres especiales mal interpretados
                nombre = re.sub(r'[^A-ZÁÉÍÓÚÜÑ\s]', '', nombre.upper())
                nombre = re.sub(r'\s+', ' ', nombre).strip()
                
                # Separar tipo y documento
                if '-' in tipo_doc_completo:
                    tipo, documento = tipo_doc_completo.split('-', 1)
                else:
                    tipo = ""
                    documento = tipo_doc_completo
                
                tipos.append(tipo)
                documentos.append(documento)
                nombres.append(nombre)

    # Si no se encontró ningún dato, lanzar una excepción específica
    if len(documentos) == 0:
        raise ValueError("No se pudo extraer el listado. El formato del PDF puede estar mal elaborado o no seguir la estructura esperada.")

    return tipos, documentos, nombres

def llenar_plantilla_con_datos(tipos, documentos, nombres, archivo_salida):
    """
    Llena la plantilla generada por plantilla.py con los datos extraídos
    SIN perder las validaciones
    """
    # Generar la plantilla usando el componente plantilla.py
    archivo_plantilla_temp = generar_plantilla()
    
    # Cargar el workbook existente
    wb = load_workbook(archivo_plantilla_temp)
    ws = wb.active
    
    # Llenar los datos en las celdas correspondientes
    for i, (tipo, documento, nombre) in enumerate(zip(tipos, documentos, nombres), start=2):
        ws[f'A{i}'] = tipo
        ws[f'B{i}'] = documento
        ws[f'C{i}'] = nombre
        # Las columnas D, E, F (DIA, MES, AÑO) se dejan vacías
    
    # Guardar el archivo final
    wb.save(archivo_salida)
    
    # Eliminar el archivo temporal
    try:
        os.unlink(archivo_plantilla_temp)
    except:
        pass

def procesar_carpeta_principal_api(ruta_carpeta):
    """
    Función principal adaptada para API que procesa todos los PDFs en las subcarpetas
    Retorna un diccionario con los resultados para JSON
    """
    resultado = {
        "procesados_exitosos": 0,
        "procesados_con_error": 0,
        "total_pdfs_encontrados": 0,
        "archivos_procesados": [],
        "errores": []
    }
    
    try:
        carpeta_principal = Path(ruta_carpeta)
        
        if not carpeta_principal.exists() or not carpeta_principal.is_dir():
            raise ValueError("La carpeta no existe o la ruta es incorrecta")
        
        # Buscar todos los PDFs en las subcarpetas
        pdfs_encontrados = buscar_pdfs_en_carpeta(carpeta_principal)
        resultado["total_pdfs_encontrados"] = len(pdfs_encontrados)
        
        if not pdfs_encontrados:
            resultado["message"] = "No se encontraron archivos PDF en las subcarpetas."
            return resultado
        
        # Procesar cada PDF encontrado
        for pdf_info in pdfs_encontrados:
            try:
                # Extraer datos del PDF
                tipos, documentos, nombres = extraer_datos_del_pdf(pdf_info['ruta'])
                
                # Definir el nombre del archivo de salida - SIEMPRE "plantilla.xlsx"
                archivo_salida = pdf_info['subcarpeta'] / "plantilla.xlsx"
                
                # Si el archivo ya existe, sobrescribir automáticamente para API
                # (en API no podemos preguntar interactivamente)
                
                # Llenar la plantilla con los datos extraídos
                llenar_plantilla_con_datos(tipos, documentos, nombres, archivo_salida)
                
                resultado["procesados_exitosos"] += 1
                resultado["archivos_procesados"].append({
                    "pdf": pdf_info['nombre_archivo'],
                    "subcarpeta": str(pdf_info['subcarpeta']),
                    "excel_generado": str(archivo_salida),
                    "registros_procesados": len(documentos)
                })
                
            except Exception as e:
                resultado["procesados_con_error"] += 1
                resultado["errores"].append({
                    "pdf": pdf_info['nombre_archivo'],
                    "subcarpeta": str(pdf_info['subcarpeta']),
                    "error": str(e)
                })
                continue
        
        # Mensaje final
        if resultado["procesados_exitosos"] > 0:
            resultado["message"] = f"Proceso completado. {resultado['procesados_exitosos']} PDFs procesados exitosamente."
        else:
            resultado["message"] = "No se pudo procesar ningún PDF."
        
        return resultado
        
    except Exception as e:
        resultado["error"] = str(e)
        return resultado

# Función para uso desde línea de comandos (manteniendo compatibilidad)
def procesar_carpeta_principal_cli():
    """
    Función original para uso desde línea de comandos
    """
    print("=" * 60)
    print("      EXTRACTOR MASIVO DE DATOS PDF A PLANTILLA EXCEL")
    print("=" * 60)
    
    try:
        # Solicitar la ruta de la carpeta principal
        ruta_carpeta = input("Por favor, ingresa la ruta completa de la carpeta principal: ").strip()
        ruta_carpeta = ruta_carpeta.strip('"\'')
        
        if not os.path.exists(ruta_carpeta) or not os.path.isdir(ruta_carpeta):
            print("❌ Error: La carpeta no existe o la ruta es incorrecta.")
            return
        
        resultado = procesar_carpeta_principal_api(ruta_carpeta)
        
        # Mostrar resultados en formato CLI
        print(f"\n📋 Total de PDFs encontrados: {resultado['total_pdfs_encontrados']}")
        
        for archivo in resultado["archivos_procesados"]:
            print(f"✅ {archivo['pdf']} -> {archivo['excel_generado']} ({archivo['registros_procesados']} registros)")
        
        for error in resultado["errores"]:
            print(f"❌ {error['pdf']}: {error['error']}")
        
        print(f"\n📊 RESUMEN:")
        print(f"✅ PDFs procesados exitosamente: {resultado['procesados_exitosos']}")
        print(f"❌ PDFs con errores: {resultado['procesados_con_error']}")
        
        if resultado["procesados_exitosos"] > 0:
            print(f"\n📝 Los archivos Excel se guardaron en sus respectivas subcarpetas")
            print("   con el nombre: plantilla.xlsx")
        
    except Exception as e:
        print(f"\n❌ Error durante el proceso: {str(e)}")
    
    input("\nPresiona Enter para salir...")

# Punto de entrada para uso desde API
if __name__ == "__main__":
    # Si se ejecuta desde línea de comandos sin argumentos, usar modo interactivo
    if len(sys.argv) == 1:
        procesar_carpeta_principal_cli()
    # Si se pasa un argumento (ruta de carpeta), usar modo API
    elif len(sys.argv) == 2:
        ruta_carpeta = sys.argv[1]
        resultado = procesar_carpeta_principal_api(ruta_carpeta)
        print(json.dumps(resultado, ensure_ascii=False, indent=2))
    else:
        print("Uso: python ExtraerInfAprendiz.py [ruta_carpeta]")
        print("Si no se proporciona ruta, se ejecuta en modo interactivo")